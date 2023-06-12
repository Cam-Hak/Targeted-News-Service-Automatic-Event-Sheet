from openpyxl import Workbook, load_workbook
from datetime import date
import os

wb = load_workbook('/Users/cameronhakenson/triage_transfer.xlsx') # Location of excel file
ws = wb.active

def restart_sheet():
    ws['A1'] = '' ; ws['A2'] = '' ; ws['A3'] = '' ; ws['A4'] = ''
    ws['A7'] = '' ; ws['A8'] = ''
    ws['A9'] = '' ; ws['A10'] = '' ; ws['A11'] = '' ; ws['A12'] = ''
    ws['A13'] = '' ; ws['A14'] = '' ; ws['A15'] = '' ; ws['A16'] = ''
    ws['A17'] = '' ; ws['A18'] = '' ; ws['A19'] = '' ; ws['A20'] = ''
    ws['A21'] = '' ; ws['A22'] = '' ; ws['A23'] = '' ; ws['A24'] = '' ; ws['A25'] = ''

def url():
    url_id = input('Enter URL ID: ')
    ws['A5'] = url_id
    full_url = input('Enter Full URL: ')
    ws['A6'] = full_url

def date_made():
    today = date.today()
    full_date = f"{today.month}/{today.day}/23"
    ws['A1'] = f'CAMERON {full_date}'
    ws['A12'] = 'Overview from the organization:'
    ws['A23'] = '* * *' ; ws['A25'] = '* * *'
    ws['A26'] = 'T41-Cameron'

def event_check():
    event_check = input('Webinar or Conference (web or con): ')
    if event_check == 'web':
        return True
    elif event_check == 'con':
        return False

def webinar_date(month, day, year):
    time = input('Enter the time: ')
    if len(month) > 3:
        ws['A9'] = f'Date of event: {month} {day}, {year}, {time}'
        date = f'Webinar Scheduled for {month} {day}, {year}'
    else:
        ws['A9'] = f'Date of event: {month}. {day}, {year}, {time}'
        date = f'Webinar Scheduled for {month}. {day}, {year}'
    return date

def conference_date(month, day, year):
    if len(month) > 3:
        ws['A9'] = f'Date of event: {month} {day}, {year}'
        date = f'Conference Scheduled for {month} {day}, {year}'
    else:
        ws['A9'] = f'Date of event: {month}. {day}, {year}'
        date = f'Conference Scheduled for {month}. {day}, {year}'
    return date

def title_section(date):
    kept_lower = ['and', 'of', 'the', 'as', 'for', 'or', 'if']
    title = input('Enter the title: ')
    title_format = title.split()
    formatted = ''
    for words in title_format:
        if words != words.upper() and '\'' not in words and words not in kept_lower:
                formatted += words.title() + " "
        elif '\'' in words:
            first_letter = words[0].upper()
            formatted += first_letter + words[1:] + " "
        elif words in kept_lower:
            formatted += words + " "
        else:
            formatted += words + " "
    title_and_date = formatted + date
    ws['A8'] = f'Title: {formatted}'
    ws['A7'] = title_and_date

def con_location():
    location = input('Enter location: ')
    ws['A10'] = 'Location of event:'
    ws['A11'] = location

def web_description():
    os.system("open /Users/cameronhakenson/triage_description.txt") # Location of description txt file
    blank_read = input('Press Enter to continue...')
    with open('triage_description.txt') as file_description:
        ws['A14'] = file_description.read()
    with open('triage_description.txt', 'w') as file_description:
        file_description.write('')
    ws['A15'] = ""

def con_agenda():
    os.system("open /Users/cameronhakenson/triage_agenda.txt") # Location of agenda txt file
    blank_read = input('Press Enter to continue...')
    with open('triage_agenda.txt') as file_agenda:
        ws['A15'] = 'Agenda: ' + file_agenda.read()
    with open('triage_agenda.txt', 'w') as file_agenda:
        file_agenda.write('')
    ws['a14'] = ''
    link_agenda = input('Enter agenda link: ')
    ws['A19'] = f'Link to full agenda: {link_agenda}'

def speakers():
    speakers = input('Enter speakers: ')
    if speakers == "" or speakers == " ":
        ws['A20'] = ""
    else:
        ws['A20'] = f'Speakers: {speakers}'

def sponsors():
    sponsors = input('Enter sponsors: ')
    if sponsors == "" or sponsors == " ":
        ws['A21'] = ""
        ws['A22'] = ""
    else:
        ws['A21'] = '* * *'
        ws['A22'] = f'Sponsors: {sponsors}'

def register():
    register = input('Enter registration link: ')
    ws['A24'] = f'Registration: {register}'
